#!/usr/bin/env python3
r"""
YOLO 分割检测框与外圈椭圆评估工具。

平时只需修改文件顶部 ``EvaluationConfig`` 配置区，然后直接运行：

    python unit_test/yolo_mask_bbox_error.py

脚本保留原有的七项检测框像素误差，并新增：

1. 只对 cls0 最外圈运行项目同款稳健 Mask 椭圆拟合。
2. 计算最外圈标签多边形与填充拟合椭圆的像素 IoU。
3. 按拟合椭圆长轴 <77px、77–311px、>311px 分三档，并保留总体统计。
4. 检测和椭圆 IoU 分目录保存，优先输出 Excel，环境不支持时回退 TXT。

The original seven bounding-box errors are:

The seven reported errors are absolute pixel errors:

    left_edge_error   = abs(pred_x1 - gt_x1)
    top_edge_error    = abs(pred_y1 - gt_y1)
    right_edge_error  = abs(pred_x2 - gt_x2)
    bottom_edge_error = abs(pred_y2 - gt_y2)

    x_axis_mean_error = (left_edge_error + right_edge_error) / 2
    y_axis_mean_error = (top_edge_error + bottom_edge_error) / 2

The ground-truth rectangle decides which size is used for the long-side length
error:

    horizontal long side:
        long_side_length_error = abs(pred_width - gt_width)

    vertical long side:
        long_side_length_error = abs(pred_height - gt_height)

By default, an error is considered qualified when it is less than or equal to
3 pixels. The threshold can be changed with ``--threshold``. The output
includes the qualification result for each error, each error's qualification
rate, the 95th percentile (P95), and the rate at which all seven errors qualify
in the same image. When visualization is enabled, images for which any one of
the seven errors exceeds the threshold are also saved to a separate directory.
The ten images with the largest per-image error are saved in one folder.  The
ranking score is the maximum of the seven errors for that image.

Expected label format (one target per line):

    class_id x1 y1 x2 y2 x3 y3 ...

Polygon coordinates must be normalized to [0, 1], as in an Ultralytics YOLO
segmentation dataset.

By default, all label classes and all prediction classes participate in
automatic matching. Every ground-truth box is compared with every prediction
box, and the pair with the highest bounding-box IoU is selected. IoU ties are
resolved using prediction confidence and then ground-truth box area. The
selected ground-truth class, prediction class, and IoU are used in the
per-image result and visualization. ``--class-id`` and
``--match-pred-class`` remain available when a restricted evaluation is
explicitly required.

Dependencies:

    pip install ultralytics opencv-python numpy
"""

from __future__ import annotations

import argparse
import copy
import csv
import statistics
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import cv2
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    # Excel 只是统计输出的增强能力。服务器没有 openpyxl 时不能影响推理和评估，
    # 主流程会自动回退为 UTF-8-BOM TXT。
    Workbook = None  # type: ignore[assignment]
    Alignment = Font = PatternFill = None  # type: ignore[assignment]
    HAS_OPENPYXL = False

try:
    from ultralytics import YOLO
except ImportError as exc:
    raise SystemExit(
        "未安装 ultralytics。请先执行：pip install ultralytics opencv-python numpy"
    ) from exc

try:
    # 直接复用项目正式版外圈拟合，避免评估脚本和板端算法各维护一套。
    from final_version import (
        Config as ProjectEllipseConfig,
        Detection as ProjectDetection,
        EllipseResult as ProjectEllipseResult,
        best_ellipse,
    )
except ImportError as exc:
    raise SystemExit(
        "无法加载 unit_test/final_version.py 的椭圆拟合模块；"
        "请确认两个脚本位于同一目录，且 requirements_pt.txt 依赖已安装。"
    ) from exc


IMAGE_SUFFIXES = {
    ".jpg",
    ".jpeg",
    ".png",
    ".bmp",
    ".tif",
    ".tiff",
    ".webp",
}

# =============================================================================
# 配置区：平时只需修改这里，然后直接运行脚本。
# 命令行参数仍然保留，用于服务器批处理时临时覆盖。
# =============================================================================
@dataclass
class EvaluationConfig:
    # 路径前的 r 可避免 Windows 反斜杠被当作转义符。
    model_path: Path = Path(r"D:\work\验证数据集\best.pt")
    images_path: Path = Path(r"D:\work\验证数据集\全部图像文件")
    labels_path: Path = Path(r"D:\work\验证数据集\内外圈数据集标签")
    output_path: Path = Path(r"D:\work\验证数据集\误差评估结果")

    # YOLO 推理。单个整数表示正方形；(height, width) 表示矩形。
    confidence: float = 0.25
    image_size: int | tuple[int, int] = 640
    device: str | None = None

    # 原有检测框误差统计。
    error_threshold_px: float = 3.0
    requested_label_class: int | None = None
    ignore_prediction_class: bool = True
    selection: str = "iou"

    # 新增外圈椭圆评估。项目当前约定 cls0 是最外圈。
    outer_class_id: int = 0
    ellipse_force_box: bool = False
    mask_threshold: float = 0.50
    # OpenCV RotatedRect 的完整长轴直径（max(width, height)），不是半长轴。
    small_major_axis_px: float = 77.0
    large_major_axis_px: float = 311.0

    # 输出。默认开启全部可视化，无需再传 --save-vis。
    recursive: bool = False
    save_visualizations: bool = True
    top_k: int = 50
    lowest_iou_count: int = 50


CFG = EvaluationConfig()

# 外圈拟合使用 final_version.py 的稳健 Mask 椭圆管线。
# 如需调整 RANSAC、Sampson LM、部分弧门限，可在这里修改对应字段。
ELLIPSE_FIT_CFG = ProjectEllipseConfig()
ELLIPSE_FIT_CFG.enable_edge_fallback = False  # 外圈只用 Mask，质量不足时走 Box 保底。

METRIC_INFO = [
    ("left_edge_error_px", "Left edge error (x1)", "left_edge_error"),
    ("top_edge_error_px", "Top edge error (y1)", "top_edge_error"),
    ("right_edge_error_px", "Right edge error (x2)", "right_edge_error"),
    ("bottom_edge_error_px", "Bottom edge error (y2)", "bottom_edge_error"),
    (
        "x_axis_mean_error_px",
        "X-axis mean error: (Left + Right) / 2",
        "x_axis_mean_error",
    ),
    (
        "y_axis_mean_error_px",
        "Y-axis mean error: (Top + Bottom) / 2",
        "y_axis_mean_error",
    ),
    (
        "long_side_length_error_px",
        "Long-side length error",
        "long_side_length_error",
    ),
]

ERROR_COLUMNS = [metric for metric, _, _ in METRIC_INFO]
METRIC_DISPLAY_NAMES = {
    metric: display_name for metric, display_name, _ in METRIC_INFO
}
METRIC_FOLDER_NAMES = {
    metric: folder_name for metric, _, folder_name in METRIC_INFO
}
PASS_COLUMNS = [
    metric[:-3] + "_within_threshold"
    for metric in ERROR_COLUMNS
]

CSV_COLUMNS = [
    "image",
    "label",
    "status",
    "message",
    "gt_class",
    "pred_class",
    "match_mode",
    "gt_target_count",
    "confidence",
    "prediction_count",
    "candidate_pair_count",
    "gt_x1",
    "gt_y1",
    "gt_x2",
    "gt_y2",
    "pred_x1",
    "pred_y1",
    "pred_x2",
    "pred_y2",
    "gt_width",
    "gt_height",
    "pred_width",
    "pred_height",
    "gt_long_axis",
    *ERROR_COLUMNS,
    "threshold_px",
    *PASS_COLUMNS,
    "all_metrics_within_threshold",
    "iou",
    # 最外圈标签 vs 项目拟合椭圆。
    "outer_ellipse_status",
    "outer_ellipse_message",
    "outer_gt_class",
    "outer_pred_class",
    "outer_confidence",
    "outer_match_box_iou",
    "outer_ellipse_source",
    "outer_ellipse_valid",
    "outer_ellipse_center_x",
    "outer_ellipse_center_y",
    "outer_ellipse_width_px",
    "outer_ellipse_height_px",
    "outer_ellipse_major_axis_px",
    "outer_ellipse_minor_axis_px",
    "outer_ellipse_angle_deg",
    "outer_size_group",
    "outer_label_area_px",
    "outer_ellipse_area_px",
    "outer_intersection_area_px",
    "outer_union_area_px",
    "outer_label_ellipse_iou",
    "outer_label_ellipse_iou_percent",
]


def parse_args(config: EvaluationConfig = CFG) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="计算标签掩码外接矩形与 YOLO 检测框之间的七项误差。"
    )
    parser.add_argument(
        "--model",
        type=Path,
        default=config.model_path,
        help=f"YOLO .pt 模型路径；配置区：{config.model_path}",
    )
    parser.add_argument(
        "--images",
        type=Path,
        default=config.images_path,
        help=f"单张图片路径或图片目录；配置区：{config.images_path}",
    )
    parser.add_argument(
        "--labels",
        type=Path,
        default=config.labels_path,
        help=f"YOLO 分割标签目录；配置区：{config.labels_path}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=config.output_path,
        help=f"结果输出目录；配置区：{config.output_path}",
    )
    parser.add_argument("--conf", type=float, default=config.confidence,
                        help="置信度阈值")
    parser.add_argument(
        "--imgsz",
        type=int,
        nargs="+",
        default=(
            [config.image_size]
            if isinstance(config.image_size, int)
            else list(config.image_size)
        ),
        metavar="N",
        help="模型推理尺寸：--imgsz 1024 或 --imgsz 768 1024（H W）",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=config.error_threshold_px,
        help="误差达标阈值，单位为像素；默认 3 px，误差小于等于阈值即达标",
    )
    parser.add_argument(
        "--device",
        default=config.device,
        help='推理设备，例如 "0"、"cpu"；不填时由 Ultralytics 自动选择',
    )
    parser.add_argument(
        "--class-id",
        type=int,
        default=config.requested_label_class,
        help="可选：只限制标签类别；默认不限制类别，自动按最大 IoU 匹配",
    )
    class_matching_group = parser.add_mutually_exclusive_group()
    class_matching_group.add_argument(
        "--ignore-pred-class",
        dest="ignore_pred_class",
        action="store_true",
        help="预测框匹配时不要求与标签同类别（默认行为）",
    )
    class_matching_group.add_argument(
        "--match-pred-class",
        dest="ignore_pred_class",
        action="store_false",
        help="只允许标签与相同类别的预测框配对",
    )
    parser.set_defaults(ignore_pred_class=config.ignore_prediction_class)
    parser.add_argument(
        "--selection",
        choices=("confidence", "iou"),
        default=config.selection,
        help="候选组合排序方式；默认按最大 IoU，confidence 仅用于兼容旧命令",
    )
    parser.add_argument(
        "--recursive",
        action="store_true",
        default=config.recursive,
        help="递归搜索图片目录，并在标签目录中保持相同的子目录结构",
    )
    visualization_group = parser.add_mutually_exclusive_group()
    visualization_group.add_argument(
        "--save-vis", dest="save_vis", action="store_true",
        help="保存检测与外圈椭圆 IoU 的逐图可视化",
    )
    visualization_group.add_argument(
        "--no-save-vis", dest="save_vis", action="store_false",
        help="仅输出 Excel/TXT 统计和 PNG 看板，不保存逐图可视化",
    )
    parser.set_defaults(save_vis=config.save_visualizations)
    parser.add_argument(
        "--top-k",
        type=int,
        default=config.top_k,
        help="按单图七项误差最大值保存多少张总体 Top 图片；默认 10",
    )
    parser.add_argument(
        "--lowest-iou-count",
        type=int,
        default=config.lowest_iou_count,
        help="保存多少张最低外圈椭圆 IoU 图片；默认 50",
    )
    args = parser.parse_args()

    if len(args.imgsz) not in (1, 2) or any(value <= 0 for value in args.imgsz):
        parser.error("--imgsz 只能提供一个正整数，或按 H W 提供两个正整数")
    # Ultralytics 对正方形接受 int，对矩形接受 (height, width)。
    args.imgsz = (args.imgsz[0] if len(args.imgsz) == 1
                  else (args.imgsz[0], args.imgsz[1]))
    if not 0.0 <= args.conf <= 1.0:
        parser.error("--conf 必须位于 0 到 1 之间")
    if args.threshold < 0.0:
        parser.error("--threshold 不能小于 0")
    if args.top_k <= 0:
        parser.error("--top-k 必须大于 0")
    if args.lowest_iou_count <= 0:
        parser.error("--lowest-iou-count 必须大于 0")
    if not 0.0 <= config.mask_threshold <= 1.0:
        parser.error("配置区 mask_threshold 必须位于 0 到 1 之间")
    if config.small_major_axis_px <= 0.0:
        parser.error("配置区 small_major_axis_px 必须大于 0")
    if config.large_major_axis_px <= config.small_major_axis_px:
        parser.error("large_major_axis_px 必须大于 small_major_axis_px")
    if config.outer_class_id < 0:
        parser.error("outer_class_id 不能小于 0")
    if config.selection not in ("confidence", "iou"):
        parser.error("selection 只能是 confidence 或 iou")
    args.outer_class_id = config.outer_class_id
    args.ellipse_force_box = config.ellipse_force_box
    args.mask_threshold = config.mask_threshold
    args.small_major_axis_px = config.small_major_axis_px
    args.large_major_axis_px = config.large_major_axis_px
    return args


def read_image(path: Path) -> np.ndarray | None:
    """Read images reliably even when a Windows path contains Chinese text."""
    try:
        data = np.fromfile(str(path), dtype=np.uint8)
    except OSError:
        return None
    if data.size == 0:
        return None
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def write_image(path: Path, image: np.ndarray) -> None:
    """Write images reliably even when a Windows path contains Chinese text."""
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower() or ".jpg"
    ok, encoded = cv2.imencode(suffix, image)
    if not ok:
        raise RuntimeError(f"无法编码可视化图片：{path}")
    encoded.tofile(str(path))


def list_images(images_path: Path, recursive: bool) -> list[Path]:
    if images_path.is_file():
        if images_path.suffix.lower() not in IMAGE_SUFFIXES:
            raise ValueError(f"不是支持的图片格式：{images_path}")
        return [images_path]
    if not images_path.is_dir():
        raise FileNotFoundError(f"图片路径不存在：{images_path}")

    iterator: Iterable[Path]
    iterator = images_path.rglob("*") if recursive else images_path.iterdir()
    return sorted(
        path
        for path in iterator
        if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES
    )


def label_path_for_image(
    image_path: Path,
    images_path: Path,
    labels_path: Path,
) -> Path:
    if images_path.is_file():
        relative = Path(image_path.name)
    else:
        relative = image_path.relative_to(images_path)
    return labels_path / relative.with_suffix(".txt")


def clip_bbox(
    bbox: tuple[float, float, float, float],
    image_width: int,
    image_height: int,
) -> tuple[float, float, float, float]:
    x1, y1, x2, y2 = bbox
    return (
        float(np.clip(x1, 0.0, float(image_width))),
        float(np.clip(y1, 0.0, float(image_height))),
        float(np.clip(x2, 0.0, float(image_width))),
        float(np.clip(y2, 0.0, float(image_height))),
    )


def read_segmentation_labels(
    label_path: Path,
    image_width: int,
    image_height: int,
    requested_class: int | None,
) -> list[tuple[int, tuple[float, float, float, float], np.ndarray]]:
    if not label_path.is_file():
        raise FileNotFoundError(f"标签文件不存在：{label_path}")

    nonempty_lines = [
        line.strip()
        for line in label_path.read_text(encoding="utf-8-sig").splitlines()
        if line.strip()
    ]
    if not nonempty_lines:
        raise ValueError(f"标签文件为空：{label_path}")

    valid_targets: list[
        tuple[int, tuple[float, float, float, float], np.ndarray]
    ] = []
    for line_number, line in enumerate(nonempty_lines, start=1):
        try:
            values = [float(value) for value in line.split()]
        except ValueError as exc:
            raise ValueError(
                f"{label_path} 第 {line_number} 行包含非数字内容"
            ) from exc

        if len(values) < 7 or (len(values) - 1) % 2 != 0:
            raise ValueError(
                f"{label_path} 第 {line_number} 行不是合法的 YOLO 分割标签"
            )

        class_id = int(values[0])
        if requested_class is not None and class_id != requested_class:
            continue

        normalized = np.asarray(values[1:], dtype=np.float64).reshape(-1, 2)
        if np.any(normalized < -1e-6) or np.any(normalized > 1.0 + 1e-6):
            raise ValueError(
                f"{label_path} 第 {line_number} 行坐标不在归一化范围 [0, 1]"
            )
        normalized = np.clip(normalized, 0.0, 1.0)

        polygon = normalized.copy()
        polygon[:, 0] *= image_width
        polygon[:, 1] *= image_height

        bbox = (
            float(np.min(polygon[:, 0])),
            float(np.min(polygon[:, 1])),
            float(np.max(polygon[:, 0])),
            float(np.max(polygon[:, 1])),
        )
        valid_targets.append((class_id, bbox, polygon))

    if not valid_targets:
        if requested_class is None:
            raise ValueError(f"标签中没有有效目标：{label_path}")
        raise ValueError(
            f"标签中不存在类别 {requested_class}：{label_path}"
        )

    return valid_targets


def bbox_area(bbox: tuple[float, float, float, float]) -> float:
    x1, y1, x2, y2 = bbox
    return max(0.0, x2 - x1) * max(0.0, y2 - y1)


def largest_ground_truth(
    targets: list[
        tuple[int, tuple[float, float, float, float], np.ndarray]
    ],
) -> tuple[int, tuple[float, float, float, float], np.ndarray]:
    """Choose a deterministic fallback target when no prediction is available."""
    return max(targets, key=lambda target: bbox_area(target[1]))


def box_iou(
    box_a: tuple[float, float, float, float],
    box_b: tuple[float, float, float, float],
) -> float:
    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b
    inter_width = max(0.0, min(ax2, bx2) - max(ax1, bx1))
    inter_height = max(0.0, min(ay2, by2) - max(ay1, by1))
    intersection = inter_width * inter_height
    area_a = max(0.0, ax2 - ax1) * max(0.0, ay2 - ay1)
    area_b = max(0.0, bx2 - bx1) * max(0.0, by2 - by1)
    union = area_a + area_b - intersection
    return intersection / union if union > 0.0 else 0.0


def choose_best_match(
    result: Any,
    gt_targets: list[
        tuple[int, tuple[float, float, float, float], np.ndarray]
    ],
    ignore_pred_class: bool,
    selection: str,
    image_width: int,
    image_height: int,
) -> tuple[
    int,
    tuple[float, float, float, float],
    np.ndarray,
    tuple[float, float, float, float] | None,
    int | None,
    float | None,
    int,
    int,
    float | None,
]:
    fallback_class, fallback_bbox, fallback_polygon = largest_ground_truth(
        gt_targets
    )
    boxes = result.boxes
    if boxes is None or len(boxes) == 0:
        return (
            fallback_class,
            fallback_bbox,
            fallback_polygon,
            None,
            None,
            None,
            0,
            0,
            None,
        )

    xyxy = boxes.xyxy.detach().cpu().numpy()
    confidences = boxes.conf.detach().cpu().numpy()
    classes = boxes.cls.detach().cpu().numpy().astype(int)
    prediction_count = len(xyxy)

    predictions: list[
        tuple[tuple[float, float, float, float], int, float]
    ] = []
    for raw_box, pred_class, confidence in zip(xyxy, classes, confidences):
        bbox = clip_bbox(
            tuple(float(value) for value in raw_box),
            image_width,
            image_height,
        )
        predictions.append((bbox, int(pred_class), float(confidence)))

    candidate_pairs: list[
        tuple[
            float,
            float,
            float,
            int,
            tuple[float, float, float, float],
            np.ndarray,
            tuple[float, float, float, float],
            int,
        ]
    ] = []
    for gt_class, gt_bbox, polygon in gt_targets:
        for pred_bbox, pred_class, confidence in predictions:
            if not ignore_pred_class and pred_class != gt_class:
                continue
            candidate_pairs.append(
                (
                    box_iou(gt_bbox, pred_bbox),
                    confidence,
                    bbox_area(gt_bbox),
                    gt_class,
                    gt_bbox,
                    polygon,
                    pred_bbox,
                    pred_class,
                )
            )

    if selection == "iou":
        selected = max(
            candidate_pairs,
            key=lambda item: (item[0], item[1], item[2]),
        ) if candidate_pairs else None
    else:
        selected = max(
            candidate_pairs,
            key=lambda item: (item[1], item[0], item[2]),
        ) if candidate_pairs else None

    if selected is None:
        return (
            fallback_class,
            fallback_bbox,
            fallback_polygon,
            None,
            None,
            None,
            prediction_count,
            0,
            None,
        )

    (
        matched_iou,
        confidence,
        _,
        gt_class,
        gt_bbox,
        polygon,
        pred_bbox,
        pred_class,
    ) = selected
    return (
        gt_class,
        gt_bbox,
        polygon,
        pred_bbox,
        pred_class,
        confidence,
        prediction_count,
        len(candidate_pairs),
        matched_iou,
    )


def choose_outer_ring_match(
    result: Any,
    gt_targets: list[
        tuple[int, tuple[float, float, float, float], np.ndarray]
    ],
    outer_class_id: int,
    image_width: int,
    image_height: int,
) -> dict[str, Any]:
    """为椭圆评估独立选择 cls0 最外圈。

    原有七项检测框误差可继续评估所有类别；这里只在
    ``outer_class_id`` 标签与同类预测之间按框 IoU 匹配，防止
    中圈或内孔被误用于外圈椭圆统计。
    """
    outer_targets = [target for target in gt_targets if target[0] == outer_class_id]
    if not outer_targets:
        return {"status": "no_outer_label"}
    if result.boxes is None or len(result.boxes) == 0:
        return {"status": "no_outer_detection"}

    xyxy = result.boxes.xyxy.detach().cpu().numpy()
    confidences = result.boxes.conf.detach().cpu().numpy()
    classes = result.boxes.cls.detach().cpu().numpy().astype(int)
    candidates: list[tuple[Any, ...]] = []
    for prediction_index, (raw_box, pred_class, confidence) in enumerate(
        zip(xyxy, classes, confidences)
    ):
        if int(pred_class) != outer_class_id:
            continue
        pred_bbox = clip_bbox(
            tuple(float(value) for value in raw_box),
            image_width,
            image_height,
        )
        for gt_class, gt_bbox, polygon in outer_targets:
            candidates.append(
                (
                    box_iou(gt_bbox, pred_bbox),
                    float(confidence),
                    bbox_area(gt_bbox),
                    int(prediction_index),
                    gt_class,
                    gt_bbox,
                    polygon,
                    pred_bbox,
                    int(pred_class),
                )
            )
    if not candidates:
        return {"status": "no_outer_detection"}

    selected = max(candidates, key=lambda item: (item[0], item[1], item[2]))
    return {
        "status": "matched",
        "box_iou": selected[0],
        "confidence": selected[1],
        "prediction_index": selected[3],
        "gt_class": selected[4],
        "gt_bbox": selected[5],
        "polygon": selected[6],
        "pred_bbox": selected[7],
        "pred_class": selected[8],
    }


def prediction_mask_at_original_size(
    result: Any,
    prediction_index: int,
    image_width: int,
    image_height: int,
    threshold: float,
) -> tuple[np.ndarray | None, np.ndarray | None]:
    """取与预测框同索引的 Mask，并统一映射回原图。

    某些 Ultralytics 版本返回 Proto 分辨率 Mask，某些在
    ``retina_masks=True`` 时直接返回原图尺寸，因此不依赖固定缩放倍数。
    """
    if result.masks is None or prediction_index >= len(result.masks.data):
        return None, None
    probability = result.masks.data[prediction_index].detach().float().cpu().numpy()
    # Results.masks.data 在多数 Ultralytics 版本中已经是 0/1 二值图。
    # 必须在 resize 前判断，否则双线性插值制造的中间值会被
    # 误当作真实软 Mask，并影响亚像素法向修正。
    has_soft_probability = bool(np.any(
        (probability > 1e-4) & (probability < 1.0 - 1e-4)))
    if probability.shape != (image_height, image_width):
        probability = cv2.resize(
            probability,
            (image_width, image_height),
            interpolation=cv2.INTER_LINEAR,
        )
    probability = np.clip(probability, 0.0, 1.0)
    binary = np.where(probability >= threshold, 255, 0).astype(np.uint8)
    probability_u8 = (
        np.rint(probability * 255.0).astype(np.uint8)
        if has_soft_probability else None
    )
    return binary, probability_u8


def major_axis_group(
    major_axis_px: float,
    small_boundary_px: float,
    large_boundary_px: float,
) -> str:
    if major_axis_px < small_boundary_px:
        return "small"
    if major_axis_px <= large_boundary_px:
        return "medium"
    return "large"


def polygon_ellipse_iou(
    polygon: np.ndarray,
    ellipse: tuple[tuple[float, float], tuple[float, float], float],
    image_width: int,
    image_height: int,
) -> dict[str, float]:
    """在原图像素坐标下计算“标签区域 vs 填充椭圆”的 IoU。"""
    label_mask = np.zeros((image_height, image_width), dtype=np.uint8)
    ellipse_mask = np.zeros_like(label_mask)
    polygon_int = np.rint(polygon).astype(np.int32).reshape(-1, 1, 2)
    cv2.fillPoly(label_mask, [polygon_int], 255)
    cv2.ellipse(ellipse_mask, ellipse, 255, cv2.FILLED, cv2.LINE_8)
    label_active = label_mask > 0
    ellipse_active = ellipse_mask > 0
    intersection = int(np.count_nonzero(label_active & ellipse_active))
    union = int(np.count_nonzero(label_active | ellipse_active))
    label_area = int(np.count_nonzero(label_active))
    ellipse_area = int(np.count_nonzero(ellipse_active))
    return {
        "iou": intersection / union if union else 0.0,
        "intersection": float(intersection),
        "union": float(union),
        "label_area": float(label_area),
        "ellipse_area": float(ellipse_area),
    }


def evaluate_outer_ellipse(
    image: np.ndarray,
    result: Any,
    gt_targets: list[
        tuple[int, tuple[float, float, float, float], np.ndarray]
    ],
    args: argparse.Namespace,
) -> dict[str, Any]:
    """运行工程同款外圈拟合，并返回逐图 IoU 与分组信息。"""
    image_height, image_width = image.shape[:2]
    matched = choose_outer_ring_match(
        result, gt_targets, args.outer_class_id, image_width, image_height)
    if matched["status"] != "matched":
        return matched

    x1, y1, x2, y2 = matched["pred_bbox"]
    x = max(0, int(np.floor(x1)))
    y = max(0, int(np.floor(y1)))
    right = min(image_width, int(np.ceil(x2)))
    bottom = min(image_height, int(np.ceil(y2)))
    if right <= x or bottom <= y:
        matched["status"] = "invalid_outer_box"
        return matched

    mask, probability = prediction_mask_at_original_size(
        result,
        matched["prediction_index"],
        image_width,
        image_height,
        args.mask_threshold,
    )
    detection = ProjectDetection(
        x=x,
        y=y,
        w=right - x,
        h=bottom - y,
        score=float(matched["confidence"]),
        class_id=int(matched["pred_class"]),
        mask=mask,
        mask_probability=probability,
    )
    fit_cfg = copy.copy(ELLIPSE_FIT_CFG)
    fitted: ProjectEllipseResult = best_ellipse(
        image,
        detection,
        fit_cfg,
        force_box=args.ellipse_force_box,
        allow_edge=False,
    )
    matched["ellipse"] = fitted.ellipse
    matched["ellipse_source"] = fitted.source
    matched["ellipse_valid"] = bool(fitted.valid)
    if not fitted.valid:
        matched["status"] = "invalid_outer_ellipse"
        return matched

    width_px, height_px = (float(value) for value in fitted.ellipse[1])
    major_axis_px = max(width_px, height_px)
    minor_axis_px = min(width_px, height_px)
    overlap = polygon_ellipse_iou(
        matched["polygon"], fitted.ellipse, image_width, image_height)
    matched.update(
        status="ok",
        ellipse_width_px=width_px,
        ellipse_height_px=height_px,
        major_axis_px=major_axis_px,
        minor_axis_px=minor_axis_px,
        size_group=major_axis_group(
            major_axis_px,
            args.small_major_axis_px,
            args.large_major_axis_px,
        ),
        **overlap,
    )
    return matched


def update_row_with_outer_ellipse(
    row: dict[str, Any],
    evaluation: dict[str, Any],
) -> None:
    """将外圈拟合结果写入逐图记录，失败时保留明确状态。"""
    row["outer_ellipse_status"] = evaluation.get("status", "not_evaluated")
    if evaluation.get("error"):
        row["outer_ellipse_message"] = str(evaluation["error"])
    if "gt_class" in evaluation:
        row["outer_gt_class"] = evaluation["gt_class"]
    if "pred_class" in evaluation:
        row["outer_pred_class"] = evaluation["pred_class"]
    if "confidence" in evaluation:
        row["outer_confidence"] = rounded(evaluation["confidence"])
    if "box_iou" in evaluation:
        row["outer_match_box_iou"] = rounded(evaluation["box_iou"])
    if "ellipse_source" in evaluation:
        row["outer_ellipse_source"] = evaluation["ellipse_source"]
    if "ellipse_valid" in evaluation:
        row["outer_ellipse_valid"] = int(bool(evaluation["ellipse_valid"]))
    ellipse = evaluation.get("ellipse")
    if ellipse is not None:
        (center_x, center_y), (width_px, height_px), angle_deg = ellipse
        row["outer_ellipse_center_x"] = rounded(center_x)
        row["outer_ellipse_center_y"] = rounded(center_y)
        row["outer_ellipse_width_px"] = rounded(width_px)
        row["outer_ellipse_height_px"] = rounded(height_px)
        row["outer_ellipse_angle_deg"] = rounded(angle_deg)
    if evaluation.get("status") != "ok":
        return
    row["outer_ellipse_major_axis_px"] = rounded(evaluation["major_axis_px"])
    row["outer_ellipse_minor_axis_px"] = rounded(evaluation["minor_axis_px"])
    row["outer_size_group"] = evaluation["size_group"]
    row["outer_label_area_px"] = rounded(evaluation["label_area"])
    row["outer_ellipse_area_px"] = rounded(evaluation["ellipse_area"])
    row["outer_intersection_area_px"] = rounded(evaluation["intersection"])
    row["outer_union_area_px"] = rounded(evaluation["union"])
    row["outer_label_ellipse_iou"] = rounded(evaluation["iou"])
    row["outer_label_ellipse_iou_percent"] = rounded(
        float(evaluation["iou"]) * 100.0)


def calculate_errors(
    gt_bbox: tuple[float, float, float, float],
    pred_bbox: tuple[float, float, float, float],
) -> dict[str, float | str]:
    gt_x1, gt_y1, gt_x2, gt_y2 = gt_bbox
    pred_x1, pred_y1, pred_x2, pred_y2 = pred_bbox

    left_edge_error = abs(pred_x1 - gt_x1)
    top_edge_error = abs(pred_y1 - gt_y1)
    right_edge_error = abs(pred_x2 - gt_x2)
    bottom_edge_error = abs(pred_y2 - gt_y2)
    x_axis_mean_error = (left_edge_error + right_edge_error) / 2.0
    y_axis_mean_error = (top_edge_error + bottom_edge_error) / 2.0

    gt_width = max(0.0, gt_x2 - gt_x1)
    gt_height = max(0.0, gt_y2 - gt_y1)
    pred_width = max(0.0, pred_x2 - pred_x1)
    pred_height = max(0.0, pred_y2 - pred_y1)

    if gt_width >= gt_height:
        long_axis = "horizontal_x"
        long_side_length_error = abs(pred_width - gt_width)
    else:
        long_axis = "vertical_y"
        long_side_length_error = abs(pred_height - gt_height)

    return {
        "gt_width": gt_width,
        "gt_height": gt_height,
        "pred_width": pred_width,
        "pred_height": pred_height,
        "gt_long_axis": long_axis,
        "left_edge_error_px": left_edge_error,
        "top_edge_error_px": top_edge_error,
        "right_edge_error_px": right_edge_error,
        "bottom_edge_error_px": bottom_edge_error,
        "x_axis_mean_error_px": x_axis_mean_error,
        "y_axis_mean_error_px": y_axis_mean_error,
        "long_side_length_error_px": long_side_length_error,
        "iou": box_iou(gt_bbox, pred_bbox),
    }


def rounded(value: float | None) -> float | str:
    return "" if value is None else round(float(value), 6)


def empty_row(image_path: Path, label_path: Path) -> dict[str, Any]:
    # Do not use ``dict_a | dict_b`` here. That syntax requires Python 3.9+,
    # while some embedded Linux boards still use Python 3.7 or 3.8.
    row = {column: "" for column in CSV_COLUMNS}
    row.update(
        {
            "image": str(image_path),
            "label": str(label_path),
        }
    )
    return row


def draw_visualization(
    image: np.ndarray,
    polygon: np.ndarray,
    gt_bbox: tuple[float, float, float, float],
    pred_bbox: tuple[float, float, float, float] | None,
    errors: dict[str, float | str] | None,
    confidence: float | None,
    gt_class: int | None = None,
    pred_class: int | None = None,
    status_text: str = "",
    threshold: float | None = None,
    outer_ellipse: dict[str, Any] | None = None,
) -> np.ndarray:
    canvas = image.copy()
    polygon_int = np.rint(polygon).astype(np.int32).reshape(-1, 1, 2)
    cv2.polylines(canvas, [polygon_int], True, (0, 255, 255), 2)

    gx1, gy1, gx2, gy2 = (int(round(value)) for value in gt_bbox)
    cv2.rectangle(canvas, (gx1, gy1), (gx2, gy2), (0, 255, 0), 2)
    cv2.putText(
        canvas,
        (
            f"Selected ground truth (class {gt_class})"
            if gt_class is not None
            else "Selected ground truth"
        ),
        (gx1, max(20, gy1 - 8)),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.6,
        (0, 255, 0),
        2,
        cv2.LINE_AA,
    )

    if pred_bbox is not None and confidence is not None:
        px1, py1, px2, py2 = (int(round(value)) for value in pred_bbox)
        cv2.rectangle(canvas, (px1, py1), (px2, py2), (0, 0, 255), 2)
        cv2.putText(
            canvas,
            (
                f"Selected prediction (class {pred_class}, conf={confidence:.3f})"
                if pred_class is not None
                else f"Selected prediction (conf={confidence:.3f})"
            ),
            (px1, min(canvas.shape[0] - 8, py1 + 22)),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.6,
            (0, 0, 255),
            2,
            cv2.LINE_AA,
        )

    if errors is not None:
        long_axis_name = (
            "X" if errors["gt_long_axis"] == "horizontal_x" else "Y"
        )
        error_items = [
            ("Left edge (x1)", float(errors["left_edge_error_px"])),
            ("Top edge (y1)", float(errors["top_edge_error_px"])),
            ("Right edge (x2)", float(errors["right_edge_error_px"])),
            ("Bottom edge (y2)", float(errors["bottom_edge_error_px"])),
            (
                "X mean (Left+Right)/2",
                float(errors["x_axis_mean_error_px"]),
            ),
            (
                "Y mean (Top+Bottom)/2",
                float(errors["y_axis_mean_error_px"]),
            ),
            (
                f"Long-side length ({long_axis_name})",
                float(errors["long_side_length_error_px"]),
            ),
        ]

        # 大图使用双栏，较小图片自动切换成单栏，防止英文说明被截断。
        panel_width = min(canvas.shape[1] - 12, 960)
        two_columns = panel_width >= 760
        panel_height = min(
            canvas.shape[0] - 12,
            218 if two_columns else 276,
        )
        if panel_width > 0 and panel_height > 0:
            overlay = canvas.copy()
            cv2.rectangle(
                overlay,
                (6, 6),
                (6 + panel_width, 6 + panel_height),
                (0, 0, 0),
                -1,
            )
            cv2.addWeighted(overlay, 0.65, canvas, 0.35, 0.0, canvas)

        cv2.putText(
            canvas,
            (
                "BOUNDING-BOX ERRORS | "
                f"AUTO-MATCH IoU={float(errors['iou']):.3f}"
            ),
            (16, 32 if two_columns else 27),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.68 if two_columns else 0.48,
            (255, 255, 255),
            2,
            cv2.LINE_AA,
        )

        if two_columns:
            text_positions = [
                (16, 68),
                (470, 68),
                (16, 102),
                (470, 102),
                (16, 136),
                (470, 136),
                (16, 170),
            ]
            metric_font_scale = 0.58
            threshold_position = (470, 170)
            legend_position = (16, 205)
        else:
            text_positions = [
                (16, 55),
                (16, 82),
                (16, 109),
                (16, 136),
                (16, 163),
                (16, 190),
                (16, 217),
            ]
            metric_font_scale = 0.43
            threshold_position = (16, 244)
            legend_position = (16, 269)

        # 绿色表示达标，红色表示超过阈值。
        for (name, value), position in zip(error_items, text_positions):
            if threshold is None:
                color = (255, 255, 255)
            elif value <= threshold:
                color = (0, 255, 0)
            else:
                color = (0, 0, 255)
            cv2.putText(
                canvas,
                f"{name}: {value:.2f} px",
                position,
                cv2.FONT_HERSHEY_SIMPLEX,
                metric_font_scale,
                color,
                2,
                cv2.LINE_AA,
            )

        if threshold is not None:
            threshold_text = f"threshold={threshold:g}px"
            cv2.putText(
                canvas,
                f"Pass <= {threshold:g} px | Fail > {threshold:g} px",
                threshold_position,
                cv2.FONT_HERSHEY_SIMPLEX,
                0.52 if two_columns else 0.40,
                (255, 255, 255),
                1,
                cv2.LINE_AA,
            )

        cv2.putText(
            canvas,
            "Yellow: selected GT mask | Green: selected GT box | "
            "Red: selected prediction",
            legend_position,
            cv2.FONT_HERSHEY_SIMPLEX,
            0.50 if two_columns else 0.36,
            (255, 255, 255),
            1,
            cv2.LINE_AA,
        )

    if status_text:
        cv2.putText(
            canvas,
            status_text,
            (10, 245 if errors is not None else 56),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.7,
            (0, 0, 255),
            2,
            cv2.LINE_AA,
        )

    # 外圈椭圆使用青色绘制；标签多边形仍为黄色。
    # 底部面板固定显示拟合来源、长轴分组和 IoU，便于甲方直接看图核对。
    if outer_ellipse is not None:
        outer_polygon = outer_ellipse.get("polygon")
        if outer_polygon is not None:
            outer_polygon_int = np.rint(outer_polygon).astype(
                np.int32).reshape(-1, 1, 2)
            polygon_overlay = canvas.copy()
            cv2.fillPoly(polygon_overlay, [outer_polygon_int], (0, 165, 255))
            cv2.addWeighted(polygon_overlay, 0.18, canvas, 0.82, 0.0, canvas)
            cv2.polylines(canvas, [outer_polygon_int], True,
                          (0, 165, 255), 3, cv2.LINE_AA)
        ellipse = outer_ellipse.get("ellipse")
        if ellipse is not None:
            cv2.ellipse(canvas, ellipse, (255, 255, 0), 3, cv2.LINE_AA)

        height, width = canvas.shape[:2]
        panel_height = min(92, max(1, height - 12))
        top = max(6, height - panel_height - 6)
        overlay = canvas.copy()
        cv2.rectangle(overlay, (6, top), (width - 6, height - 6),
                      (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.72, canvas, 0.28, 0.0, canvas)

        ellipse_status = str(outer_ellipse.get("status", "not_evaluated"))
        if ellipse_status == "ok":
            source = str(outer_ellipse.get("ellipse_source", "unknown"))
            major = float(outer_ellipse["major_axis_px"])
            minor = float(outer_ellipse["minor_axis_px"])
            size_group = str(outer_ellipse["size_group"])
            ellipse_iou = float(outer_ellipse["iou"])
            first_line = (
                f"OUTER ELLIPSE | source={source} | major={major:.2f}px | "
                f"minor={minor:.2f}px | group={size_group}"
            )
            second_line = (
                f"GT LABEL vs FITTED ELLIPSE IoU: {ellipse_iou:.4f} "
                f"({ellipse_iou * 100.0:.2f}%) | "
                f"outer-box match IoU={float(outer_ellipse['box_iou']):.4f}"
            )
            color = (0, 255, 255)
        else:
            first_line = f"OUTER ELLIPSE | status={ellipse_status}"
            second_line = (
                "Orange: outer GT polygon | Cyan: fitted outer ellipse | "
                "IoU unavailable"
            )
            color = (0, 0, 255)
        font_scale = 0.58 if width >= 1000 else 0.38
        cv2.putText(canvas, first_line, (16, top + 33),
                    cv2.FONT_HERSHEY_SIMPLEX, font_scale, color, 2,
                    cv2.LINE_AA)
        cv2.putText(canvas, second_line, (16, top + 70),
                    cv2.FONT_HERSHEY_SIMPLEX, font_scale,
                    (255, 255, 255), 1, cv2.LINE_AA)
    return canvas


def draw_outer_ellipse_iou_visualization(
    image: np.ndarray,
    evaluation: dict[str, Any],
) -> np.ndarray:
    """生成只用于椭圆 IoU 核对的干净画面。

    绿色边界是 cls0 标签外圈，红色边界是最终拟合椭圆。区域颜色用于解释 IoU：
    绿色为仅标签、红色为仅椭圆、黄色为两者交集。
    """
    canvas = image.copy()
    height, width = canvas.shape[:2]
    polygon = evaluation.get("polygon")
    ellipse = evaluation.get("ellipse")

    if polygon is not None:
        label_mask = np.zeros((height, width), dtype=np.uint8)
        polygon_int = np.rint(polygon).astype(np.int32).reshape(-1, 1, 2)
        cv2.fillPoly(label_mask, [polygon_int], 255)
    else:
        label_mask = np.zeros((height, width), dtype=np.uint8)
        polygon_int = None

    ellipse_mask = np.zeros_like(label_mask)
    if ellipse is not None:
        cv2.ellipse(ellipse_mask, ellipse, 255, cv2.FILLED, cv2.LINE_8)

    label_active = label_mask > 0
    ellipse_active = ellipse_mask > 0
    region_overlay = canvas.copy()
    region_overlay[label_active & ~ellipse_active] = (0, 190, 0)
    region_overlay[ellipse_active & ~label_active] = (0, 0, 220)
    region_overlay[label_active & ellipse_active] = (0, 215, 255)
    active = label_active | ellipse_active
    canvas[active] = cv2.addWeighted(
        region_overlay, 0.34, canvas, 0.66, 0.0)[active]

    # 两条边界使用明显不同的颜色和线宽，缩略图状态下也能区分。
    if polygon_int is not None:
        cv2.polylines(canvas, [polygon_int], True, (0, 255, 0), 3,
                      cv2.LINE_AA)
    if ellipse is not None:
        cv2.ellipse(canvas, ellipse, (0, 0, 255), 3, cv2.LINE_AA)

    panel_height = min(126, max(1, height - 12))
    top = max(6, height - panel_height - 6)
    panel = canvas.copy()
    cv2.rectangle(panel, (6, top), (width - 6, height - 6), (0, 0, 0), -1)
    cv2.addWeighted(panel, 0.76, canvas, 0.24, 0.0, canvas)
    status = str(evaluation.get("status", "not_evaluated"))
    font_scale = 0.60 if width >= 1000 else 0.40
    cv2.putText(
        canvas,
        "GREEN EDGE: OUTER GT LABEL   RED EDGE: FITTED OUTER ELLIPSE",
        (16, top + 31), cv2.FONT_HERSHEY_SIMPLEX, font_scale,
        (255, 255, 255), 2, cv2.LINE_AA)
    cv2.putText(
        canvas,
        "AREA: green=GT only, red=ellipse only, yellow=intersection",
        (16, top + 62), cv2.FONT_HERSHEY_SIMPLEX, font_scale,
        (255, 255, 255), 1, cv2.LINE_AA)
    if status == "ok":
        iou = float(evaluation["iou"])
        cv2.putText(
            canvas,
            f"IoU={iou:.4f} ({iou * 100.0:.2f}%) | "
            f"major={float(evaluation['major_axis_px']):.2f}px | "
            f"group={evaluation['size_group']} | "
            f"source={evaluation['ellipse_source']}",
            (16, top + 99), cv2.FONT_HERSHEY_SIMPLEX, font_scale,
            (0, 255, 255), 2, cv2.LINE_AA)
    else:
        cv2.putText(
            canvas, f"OUTER ELLIPSE IoU UNAVAILABLE | status={status}",
            (16, top + 99), cv2.FONT_HERSHEY_SIMPLEX, font_scale,
            (0, 0, 255), 2, cv2.LINE_AA)
    return canvas


def visualization_path_for_image(
    image_path: Path,
    images_path: Path,
    visualization_dir: Path,
) -> Path:
    if images_path.is_file():
        relative_visualization = Path(image_path.name)
    else:
        relative_visualization = image_path.relative_to(images_path)
    return visualization_dir / relative_visualization


def draw_top_k_banner(
    image: np.ndarray,
    metric_name: str,
    rank: int,
    top_k: int,
    value: float,
    source_name: str,
) -> np.ndarray:
    """Add a clear ranking banner to a Top-K visualization."""
    canvas = image.copy()
    height, width = canvas.shape[:2]
    panel_height = min(82, max(1, height - 12))
    y1 = max(6, height - panel_height - 6)
    y2 = height - 6

    overlay = canvas.copy()
    cv2.rectangle(overlay, (6, y1), (width - 6, y2), (0, 0, 0), -1)
    cv2.addWeighted(overlay, 0.72, canvas, 0.28, 0.0, canvas)

    font_scale = 0.62 if width >= 1000 else 0.42
    cv2.putText(
        canvas,
        (
            f"TOP ERROR RANK {rank:02d}/{top_k:02d} | "
            f"{metric_name}: {value:.2f} px"
        ),
        (16, min(height - 34, y1 + 31)),
        cv2.FONT_HERSHEY_SIMPLEX,
        font_scale,
        (0, 165, 255),
        2,
        cv2.LINE_AA,
    )
    cv2.putText(
        canvas,
        f"Source image: {source_name}",
        (16, min(height - 10, y1 + 62)),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.50 if width >= 1000 else 0.34,
        (255, 255, 255),
        1,
        cv2.LINE_AA,
    )
    return canvas


def per_image_max_error(row: dict[str, Any]) -> tuple[float, str]:
    """返回单张图片七项检测误差中的最大值及对应指标。"""
    metric = max(ERROR_COLUMNS, key=lambda key: float(row[key]))
    return float(row[metric]), metric


def detection_top_k_records(
    rows: list[dict[str, Any]],
    top_k: int,
    output_dir: Path,
) -> list[dict[str, Any]]:
    """生成总体最大误差排行；是否保存图片不影响统计表中的排行。"""
    ranked_rows = sorted(
        (row for row in rows if row.get("status") == "ok"),
        key=lambda row: (
            -per_image_max_error(row)[0],
            str(row["image"]).lower(),
        ),
    )[:top_k]
    records: list[dict[str, Any]] = []
    for rank, row in enumerate(ranked_rows, start=1):
        value, metric = per_image_max_error(row)
        records.append(
            {
                "rank": rank,
                "image": str(row["image"]),
                "max_error_px": value,
                "metric": METRIC_DISPLAY_NAMES[metric],
                "visualization": str(output_dir / f"rank_{rank:02d}.jpg"),
            }
        )
    return records


def write_overall_top_k_visualizations(
    output_dir: Path,
    rows: list[dict[str, Any]],
    images_path: Path,
    visualization_dir: Path,
    top_k: int,
) -> tuple[list[dict[str, Any]], int]:
    """按单图最大检测误差统一排序，不再为七项指标重复建立 Top-K。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    records = detection_top_k_records(rows, top_k, output_dir)
    saved_count = 0
    for record in records:
        rank = int(record["rank"])
        source_image = Path(str(record["image"]))
        source_visualization = visualization_path_for_image(
            source_image, images_path, visualization_dir)
        value = float(record["max_error_px"])
        destination = Path(str(record["visualization"]))
        visualized = read_image(source_visualization)
        if visualized is not None:
            ranked_visualization = draw_top_k_banner(
                visualized,
                f"Maximum of 7 errors ({record['metric']})",
                rank,
                top_k,
                value,
                source_image.name,
            )
            write_image(destination, ranked_visualization)
            saved_count += 1
    return records, saved_count


def percentile(values: list[float], value: float) -> float:
    """集中处理分位数，确保 Excel、TXT 和 PNG 使用一致定义。"""
    return float(np.percentile(np.asarray(values, dtype=np.float64), value))


def statistic_text(row: dict[str, Any], key: str, scale: float = 1.0) -> str:
    """空分组输出 --，避免把“无样本”误读成指标等于零。"""
    if int(row.get("count", 0)) == 0 or key not in row:
        return "--"
    return f"{float(row[key]) * scale:.6f}"


def detection_metric_statistics(
    rows: list[dict[str, Any]],
    threshold_px: float,
    small_boundary_px: float,
    large_boundary_px: float,
) -> list[dict[str, Any]]:
    """生成检测误差总体和长轴三档统计，供 Excel/TXT 与 PNG 共用。"""
    valid_detection_rows = [
        row for row in rows if row.get("status") == "ok"
    ]
    groupable_rows = [
        row for row in valid_detection_rows
        if row.get("outer_ellipse_status") == "ok"
    ]
    output_rows: list[dict[str, Any]] = []
    for group in ("overall", "small", "medium", "large"):
        candidates = (
            valid_detection_rows
            if group == "overall"
            else [
                row for row in groupable_rows
                if row.get("outer_size_group") == group
            ]
        )
        denominator = (
            len(valid_detection_rows)
            if group == "overall"
            else len(groupable_rows)
        )
        for metric in ERROR_COLUMNS:
            values = [float(row[metric]) for row in candidates]
            output: dict[str, Any] = {
                "size_group": group,
                "major_axis_range": ellipse_group_label(
                    group, small_boundary_px, large_boundary_px),
                "metric": METRIC_DISPLAY_NAMES[metric],
                "metric_key": metric,
                "unit": "px",
                "direction": "lower_is_better",
                "count": len(values),
                "sample_percent": (
                    len(candidates) / denominator * 100.0
                    if denominator else 0.0
                ),
                "threshold": threshold_px,
                "pass_count": 0,
                "pass_rate_percent": 0.0,
            }
            if values:
                pass_count = sum(value <= threshold_px for value in values)
                output.update(
                    mean=sum(values) / len(values),
                    std=statistics.pstdev(values) if len(values) > 1 else 0.0,
                    min=min(values),
                    p05=percentile(values, 5),
                    median=percentile(values, 50),
                    p95=percentile(values, 95),
                    max=max(values),
                    pass_count=pass_count,
                    pass_rate_percent=pass_count / len(values) * 100.0,
                )
            output_rows.append(output)
    return output_rows


def write_detection_statistics_txt(
    path: Path,
    statistics_rows: list[dict[str, Any]],
    groups: tuple[str, ...],
    threshold_px: float,
) -> None:
    """写检测总体或分组 TXT；使用制表符，Excel 也可直接分列打开。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    selected = [
        row for row in statistics_rows if row["size_group"] in groups
    ]
    with path.open("w", encoding="utf-8-sig") as file:
        file.write(f"检测误差达标阈值：<= {threshold_px:g} px\n")
        if groups != ("overall",):
            file.write(
                "说明：分组只统计同时具有有效外圈拟合长轴的检测成功样本；"
                "占比为该组在可分组样本中的比例。\n"
            )
        file.write(
            "分组\t长轴范围\t样本数\t样本占比(%)\t指标\t均值(px)\t"
            "最小值(px)\t最大值(px)\tP95(px)\t达标数\t达标率(%)\n"
        )
        for row in selected:
            file.write(
                f"{row['size_group']}\t{row['major_axis_range']}\t"
                f"{row['count']}\t{row['sample_percent']:.6f}\t"
                f"{row['metric']}\t"
                f"{statistic_text(row, 'mean')}\t"
                f"{statistic_text(row, 'min')}\t"
                f"{statistic_text(row, 'max')}\t"
                f"{statistic_text(row, 'p95')}\t"
                f"{row['pass_count']}\t"
                f"{float(row['pass_rate_percent']):.6f}\n"
            )


def write_per_image_detection_txt(
    path: Path,
    rows: list[dict[str, Any]],
) -> None:
    """每张图片固定一行，失败样本也保留状态，方便定位漏检或坏数据。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "图片", "状态", "外圈长轴分组", "外圈长轴(px)",
        *[METRIC_DISPLAY_NAMES[metric] for metric in ERROR_COLUMNS],
        "单图最大误差(px)", "最大误差指标", "七项全部达标",
    ]
    with path.open("w", encoding="utf-8-sig") as file:
        file.write("\t".join(headers) + "\n")
        for row in rows:
            valid = row.get("status") == "ok"
            max_value, max_metric = (
                per_image_max_error(row) if valid else (None, "")
            )
            values = [
                str(row.get("image", "")),
                str(row.get("status", "")),
                str(row.get("outer_size_group", "")),
                (
                    f"{float(row['outer_ellipse_major_axis_px']):.6f}"
                    if row.get("outer_ellipse_major_axis_px") not in ("", None)
                    else ""
                ),
            ]
            values.extend(
                f"{float(row[metric]):.6f}" if valid else ""
                for metric in ERROR_COLUMNS
            )
            values.extend(
                [
                    f"{max_value:.6f}" if max_value is not None else "",
                    METRIC_DISPLAY_NAMES[max_metric] if max_metric else "",
                    (
                        str(int(row["all_metrics_within_threshold"]))
                        if valid else ""
                    ),
                ]
            )
            file.write("\t".join(values) + "\n")


def style_excel_sheet(worksheet: Any) -> None:
    """统一 Excel 表头、筛选、冻结和列宽，打开即可查看。"""
    if worksheet.max_row < 1:
        return
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(60, max(10, max(len(value) for value in values) + 2))
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def append_excel_rows(
    worksheet: Any,
    headers: list[str],
    values: Iterable[Iterable[Any]],
) -> None:
    worksheet.append(headers)
    for row_values in values:
        worksheet.append(list(row_values))
    style_excel_sheet(worksheet)


def detection_statistics_excel_values(
    statistics_rows: list[dict[str, Any]],
    groups: tuple[str, ...],
) -> Iterable[list[Any]]:
    for row in statistics_rows:
        if row["size_group"] not in groups:
            continue
        yield [
            row["size_group"],
            row["major_axis_range"],
            int(row["count"]),
            float(row["sample_percent"]),
            row["metric"],
            row.get("mean"),
            row.get("min"),
            row.get("max"),
            row.get("p95"),
            float(row["threshold"]),
            int(row["pass_count"]),
            float(row["pass_rate_percent"]),
        ]


def write_detection_excel(
    path: Path,
    statistics_rows: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    top_records: list[dict[str, Any]],
) -> None:
    """将检测总体、分组、逐图和 Top10 收入同一个工作簿。"""
    if not HAS_OPENPYXL or Workbook is None:
        raise RuntimeError("openpyxl unavailable")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    overall_sheet = workbook.active
    overall_sheet.title = "总体误差"
    statistics_headers = [
        "分组", "长轴范围", "样本数", "样本占比(%)", "指标",
        "均值(px)", "最小值(px)", "最大值(px)", "P95(px)",
        "阈值(px)", "达标数", "达标率(%)",
    ]
    append_excel_rows(
        overall_sheet,
        statistics_headers,
        detection_statistics_excel_values(statistics_rows, ("overall",)),
    )
    grouped_sheet = workbook.create_sheet("分组误差")
    append_excel_rows(
        grouped_sheet,
        statistics_headers,
        detection_statistics_excel_values(
            statistics_rows, ("small", "medium", "large")),
    )

    per_image_sheet = workbook.create_sheet("逐图误差")
    per_image_headers = [
        "图片", "状态", "外圈长轴分组", "外圈长轴(px)",
        *[METRIC_DISPLAY_NAMES[metric] for metric in ERROR_COLUMNS],
        "单图最大误差(px)", "最大误差指标", "七项全部达标",
    ]
    per_image_values: list[list[Any]] = []
    for row in rows:
        valid = row.get("status") == "ok"
        max_value, max_metric = (
            per_image_max_error(row) if valid else (None, "")
        )
        per_image_values.append(
            [
                str(row.get("image", "")),
                str(row.get("status", "")),
                str(row.get("outer_size_group", "")),
                (
                    float(row["outer_ellipse_major_axis_px"])
                    if row.get("outer_ellipse_major_axis_px") not in ("", None)
                    else None
                ),
                *[
                    float(row[metric]) if valid else None
                    for metric in ERROR_COLUMNS
                ],
                max_value,
                METRIC_DISPLAY_NAMES[max_metric] if max_metric else "",
                (
                    int(row["all_metrics_within_threshold"])
                    if valid else None
                ),
            ]
        )
    append_excel_rows(per_image_sheet, per_image_headers, per_image_values)

    top_sheet = workbook.create_sheet("总体最大误差Top10")
    append_excel_rows(
        top_sheet,
        ["排名", "图片", "单图最大误差(px)", "最大误差指标", "可视化文件"],
        (
            [
                record["rank"], record["image"], record["max_error_px"],
                record["metric"], record["visualization"],
            ]
            for record in top_records
        ),
    )
    workbook.save(path)


def write_ranking_txt(
    path: Path,
    headers: list[str],
    records: Iterable[Iterable[Any]],
) -> None:
    """openpyxl 不可用时，为排行信息提供同等内容的 TXT 回退。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig") as file:
        file.write("\t".join(headers) + "\n")
        for record in records:
            file.write("\t".join(str(value) for value in record) + "\n")


def write_per_image_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_run_config(path: Path, args: argparse.Namespace) -> None:
    """保存本次实际参数，防止多轮实验后无法确认使用的是哪组配置。"""
    values = {
        "model": args.model,
        "images": args.images,
        "labels": args.labels,
        "output": args.output,
        "confidence": args.conf,
        "image_size": args.imgsz,
        "device": args.device,
        "bbox_error_threshold_px": args.threshold,
        "bbox_requested_label_class": args.class_id,
        "bbox_ignore_prediction_class": args.ignore_pred_class,
        "bbox_selection": args.selection,
        "outer_class_id": args.outer_class_id,
        "ellipse_force_box": args.ellipse_force_box,
        "mask_threshold": args.mask_threshold,
        "small_major_axis_px": args.small_major_axis_px,
        "large_major_axis_px": args.large_major_axis_px,
        "ellipse_ransac_iterations": ELLIPSE_FIT_CFG.ellipse_ransac_iters,
        "ellipse_inlier_px": ELLIPSE_FIT_CFG.ellipse_inlier_px,
        "ellipse_min_quality": ELLIPSE_FIT_CFG.ellipse_min_quality,
        "save_visualizations": args.save_vis,
        "top_k": args.top_k,
    }
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=["key", "value"])
        writer.writeheader()
        for key, value in values.items():
            writer.writerow({"key": key, "value": value})


def write_summary_csv(
    path: Path,
    rows: list[dict[str, Any]],
    threshold: float,
) -> None:
    valid_rows = [row for row in rows if row["status"] == "ok"]
    summary_columns = [
        "metric",
        "metric_key",
        "count",
        "mean_px",
        "std_px",
        "min_px",
        "max_px",
        "p95_px",
        "threshold_px",
        "pass_count",
        "pass_rate_percent",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=summary_columns)
        writer.writeheader()
        for metric in ERROR_COLUMNS:
            values = [float(row[metric]) for row in valid_rows]
            if not values:
                writer.writerow(
                    {
                        "metric": METRIC_DISPLAY_NAMES[metric],
                        "metric_key": metric,
                        "count": 0,
                        "threshold_px": threshold,
                        "pass_count": 0,
                    }
                )
                continue
            pass_count = sum(value <= threshold for value in values)
            writer.writerow(
                {
                    "metric": METRIC_DISPLAY_NAMES[metric],
                    "metric_key": metric,
                    "count": len(values),
                    # sum / len is used instead of statistics.fmean so the
                    # script remains compatible with Python 3.7.
                    "mean_px": round(sum(values) / len(values), 6),
                    "std_px": round(
                        statistics.pstdev(values) if len(values) > 1 else 0.0,
                        6,
                    ),
                    "min_px": round(min(values), 6),
                    "max_px": round(max(values), 6),
                    "p95_px": round(
                        float(np.percentile(
                            np.asarray(values, dtype=np.float64), 95
                        )),
                        6,
                    ),
                    "threshold_px": threshold,
                    "pass_count": pass_count,
                    "pass_rate_percent": round(
                        pass_count / len(values) * 100.0,
                        6,
                    ),
                }
            )

        # 将全部图片的七项误差合并，给出总体误差分布的 P95。
        combined_values = [
            float(row[metric])
            for row in valid_rows
            for metric in ERROR_COLUMNS
        ]
        if combined_values:
            combined_pass_count = sum(
                value <= threshold for value in combined_values
            )
            writer.writerow(
                {
                    "metric": "All seven error values combined",
                    "metric_key": "all_error_values_combined",
                    "count": len(combined_values),
                    "mean_px": round(
                        sum(combined_values) / len(combined_values), 6
                    ),
                    "std_px": round(
                        statistics.pstdev(combined_values)
                        if len(combined_values) > 1
                        else 0.0,
                        6,
                    ),
                    "min_px": round(min(combined_values), 6),
                    "max_px": round(max(combined_values), 6),
                    "p95_px": round(
                        float(np.percentile(
                            np.asarray(combined_values, dtype=np.float64), 95
                        )),
                        6,
                    ),
                    "threshold_px": threshold,
                    "pass_count": combined_pass_count,
                    "pass_rate_percent": round(
                        combined_pass_count / len(combined_values) * 100.0,
                        6,
                    ),
                }
            )

        all_metrics_pass_count = sum(
            int(row["all_metrics_within_threshold"]) for row in valid_rows
        )
        writer.writerow(
            {
                "metric": "All seven errors within threshold",
                "metric_key": "all_metrics_within_threshold",
                "count": len(valid_rows),
                "threshold_px": threshold,
                "pass_count": all_metrics_pass_count,
                "pass_rate_percent": (
                    round(
                        all_metrics_pass_count / len(valid_rows) * 100.0,
                        6,
                    )
                    if valid_rows
                    else ""
                ),
            }
        )


def ellipse_group_label(
    group: str,
    small_boundary_px: float,
    large_boundary_px: float,
) -> str:
    labels = {
        "overall": "Overall: all valid outer ellipses",
        "small": f"Small: major < {small_boundary_px:g}px",
        "medium": (
            f"Medium: {small_boundary_px:g}px <= major <= "
            f"{large_boundary_px:g}px"
        ),
        "large": f"Large: major > {large_boundary_px:g}px",
    }
    return labels[group]


def ellipse_iou_group_statistics(
    rows: list[dict[str, Any]],
    small_boundary_px: float,
    large_boundary_px: float,
) -> list[dict[str, Any]]:
    """按总体及拟合外圈长轴三档统计标签/椭圆 IoU。"""
    valid_rows = [
        row for row in rows if row.get("outer_ellipse_status") == "ok"
    ]
    statistics_rows: list[dict[str, Any]] = []
    for group in ("overall", "small", "medium", "large"):
        group_rows = (
            valid_rows
            if group == "overall"
            else [row for row in valid_rows if row["outer_size_group"] == group]
        )
        values = [float(row["outer_label_ellipse_iou"]) for row in group_rows]
        sources = Counter(str(row["outer_ellipse_source"]) for row in group_rows)
        output: dict[str, Any] = {
            "size_group": group,
            "major_axis_range": ellipse_group_label(
                group, small_boundary_px, large_boundary_px),
            "count": len(values),
            "mask_fit_count": sources.get("mask", 0),
            "box_fallback_count": sources.get("box", 0),
        }
        if values:
            array = np.asarray(values, dtype=np.float64)
            output.update(
                mean_iou=round(float(np.mean(array)), 6),
                mean_iou_percent=round(float(np.mean(array)) * 100.0, 6),
                std_iou=round(float(np.std(array)), 6),
                min_iou=round(float(np.min(array)), 6),
                p05_iou=round(float(np.percentile(array, 5)), 6),
                median_iou=round(float(np.median(array)), 6),
                p95_iou=round(float(np.percentile(array, 95)), 6),
                max_iou=round(float(np.max(array)), 6),
                mean_major_axis_px=round(
                    sum(float(row["outer_ellipse_major_axis_px"])
                        for row in group_rows) / len(group_rows), 6),
            )
        statistics_rows.append(output)
    return statistics_rows


def write_ellipse_iou_statistics_txt(
    path: Path,
    rows: list[dict[str, Any]],
    small_boundary_px: float,
    large_boundary_px: float,
) -> list[dict[str, Any]]:
    """保存 IoU 总体及长轴分组的占比、极值、均值和 P95。"""
    summary_rows = ellipse_iou_group_statistics(
        rows, small_boundary_px, large_boundary_px)
    valid_count = int(summary_rows[0]["count"])
    status_counts = Counter(
        str(row.get("outer_ellipse_status") or "not_evaluated")
        for row in rows
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig") as file:
        file.write(
            "分组\t长轴范围\t样本数\t有效IoU样本占比(%)\t"
            "平均IoU(%)\t最小IoU(%)\t最大IoU(%)\tP95 IoU(%)\n"
        )
        for item in summary_rows:
            count = int(item["count"])
            proportion = count / valid_count * 100.0 if valid_count else 0.0
            file.write(
                f"{item['size_group']}\t{item['major_axis_range']}\t"
                f"{count}\t{proportion:.6f}\t"
                f"{statistic_text(item, 'mean_iou', 100.0)}\t"
                f"{statistic_text(item, 'min_iou', 100.0)}\t"
                f"{statistic_text(item, 'max_iou', 100.0)}\t"
                f"{statistic_text(item, 'p95_iou', 100.0)}\n"
            )

        # 同文件保留状态占比，避免只观察拟合成功样本而忽略失败率。
        file.write("\n外圈椭圆评估状态\t数量\t占全部图片比例(%)\n")
        total_count = len(rows)
        for status, count in sorted(status_counts.items()):
            proportion = count / total_count * 100.0 if total_count else 0.0
            file.write(f"{status}\t{count}\t{proportion:.6f}\n")
    return summary_rows


def write_ellipse_iou_excel(
    path: Path,
    summary_rows: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    lowest_records: list[dict[str, Any]],
) -> None:
    """将 IoU 总体/分组、状态、逐图和最低50收入一个工作簿。"""
    if not HAS_OPENPYXL or Workbook is None:
        raise RuntimeError("openpyxl unavailable")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "总体及分组"
    valid_count = int(summary_rows[0]["count"]) if summary_rows else 0
    append_excel_rows(
        summary_sheet,
        [
            "分组", "长轴范围", "样本数", "有效IoU样本占比(%)",
            "平均IoU(%)", "最小IoU(%)", "最大IoU(%)", "P95 IoU(%)",
        ],
        (
            [
                item["size_group"],
                item["major_axis_range"],
                int(item["count"]),
                (
                    int(item["count"]) / valid_count * 100.0
                    if valid_count else 0.0
                ),
                (
                    float(item["mean_iou"]) * 100.0
                    if "mean_iou" in item else None
                ),
                (
                    float(item["min_iou"]) * 100.0
                    if "min_iou" in item else None
                ),
                (
                    float(item["max_iou"]) * 100.0
                    if "max_iou" in item else None
                ),
                (
                    float(item["p95_iou"]) * 100.0
                    if "p95_iou" in item else None
                ),
            ]
            for item in summary_rows
        ),
    )

    status_sheet = workbook.create_sheet("评估状态")
    status_counts = Counter(
        str(row.get("outer_ellipse_status") or "not_evaluated")
        for row in rows
    )
    append_excel_rows(
        status_sheet,
        ["外圈椭圆评估状态", "数量", "占全部图片比例(%)"],
        (
            [
                status,
                count,
                count / len(rows) * 100.0 if rows else 0.0,
            ]
            for status, count in sorted(status_counts.items())
        ),
    )

    per_image_sheet = workbook.create_sheet("逐图IoU")
    append_excel_rows(
        per_image_sheet,
        [
            "图片", "评估状态", "外圈长轴(px)", "分组", "拟合来源",
            "IoU", "IoU占比(%)",
        ],
        (
            [
                str(row.get("image", "")),
                str(row.get("outer_ellipse_status", "")),
                (
                    float(row["outer_ellipse_major_axis_px"])
                    if row.get("outer_ellipse_major_axis_px") not in ("", None)
                    else None
                ),
                str(row.get("outer_size_group", "")),
                str(row.get("outer_ellipse_source", "")),
                (
                    float(row["outer_label_ellipse_iou"])
                    if row.get("outer_label_ellipse_iou") not in ("", None)
                    else None
                ),
                (
                    float(row["outer_label_ellipse_iou"]) * 100.0
                    if row.get("outer_label_ellipse_iou") not in ("", None)
                    else None
                ),
            ]
            for row in rows
        ),
    )

    lowest_sheet = workbook.create_sheet("最低IoU前50")
    append_excel_rows(
        lowest_sheet,
        [
            "排名", "IoU", "IoU占比(%)", "外圈长轴(px)", "分组",
            "拟合来源", "原图片", "可视化文件",
        ],
        (
            [
                record["rank"], record["iou"],
                float(record["iou"]) * 100.0, record["major_axis_px"],
                record["size_group"], record["source"],
                record["image"], record["visualization"],
            ]
            for record in lowest_records
        ),
    )
    workbook.save(path)


def write_ellipse_iou_summary(
    path: Path,
    rows: list[dict[str, Any]],
    small_boundary_px: float,
    large_boundary_px: float,
) -> list[dict[str, Any]]:
    summary_rows = ellipse_iou_group_statistics(
        rows, small_boundary_px, large_boundary_px)
    columns = [
        "size_group", "major_axis_range", "count",
        "mask_fit_count", "box_fallback_count", "mean_major_axis_px",
        "mean_iou", "mean_iou_percent", "std_iou", "min_iou",
        "p05_iou", "median_iou", "p95_iou", "max_iou",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        writer.writerows(summary_rows)
    return summary_rows


def write_outer_ellipse_status_summary(
    path: Path,
    rows: list[dict[str, Any]],
) -> Counter[str]:
    """记录所有图片的外圈评估状态，避免只看有效样本产生幸存者偏差。"""
    counts: Counter[str] = Counter(
        str(row.get("outer_ellipse_status") or "not_evaluated")
        for row in rows
    )
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=["status", "count"])
        writer.writeheader()
        for status, count in sorted(counts.items()):
            writer.writerow({"status": status, "count": count})
    return counts


GROUPED_METRIC_INFO = [
    *[
        (metric, METRIC_DISPLAY_NAMES[metric], "px", "lower_is_better")
        for metric in ERROR_COLUMNS
    ],
    ("iou", "Selected detection-box IoU", "ratio", "higher_is_better"),
    ("outer_match_box_iou", "Outer-ring match box IoU",
     "ratio", "higher_is_better"),
    ("outer_label_ellipse_iou", "Outer label / fitted ellipse IoU",
     "ratio", "higher_is_better"),
]
DETECTION_ERROR_METRIC_INFO = GROUPED_METRIC_INFO[:len(ERROR_COLUMNS)]

METRIC_SUMMARY_COLUMNS = [
    "size_group", "major_axis_range", "metric", "metric_key",
    "unit", "direction", "count", "mean", "std", "min", "p05",
    "median", "p95", "max", "threshold", "pass_count",
    "pass_rate_percent",
]


def write_metric_summary_rows(
    path: Path,
    rows: list[dict[str, Any]],
) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=METRIC_SUMMARY_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def grouped_rows(
    valid_outer_rows: list[dict[str, Any]],
    group: str,
) -> list[dict[str, Any]]:
    if group == "overall":
        return valid_outer_rows
    return [
        row for row in valid_outer_rows
        if row.get("outer_size_group") == group
    ]


def write_all_metrics_by_outer_size(
    path: Path,
    rows: list[dict[str, Any]],
    threshold_px: float,
    small_boundary_px: float,
    large_boundary_px: float,
) -> list[dict[str, Any]]:
    """输出总体及长轴三档的全部评价指标，采用便于后处理的长表格式。"""
    valid_outer_rows = [
        row for row in rows if row.get("outer_ellipse_status") == "ok"
    ]
    output_rows: list[dict[str, Any]] = []
    for group in ("overall", "small", "medium", "large"):
        candidates = grouped_rows(valid_outer_rows, group)
        for metric, display_name, unit, direction in GROUPED_METRIC_INFO:
            values = [
                float(row[metric])
                for row in candidates
                if row.get(metric) not in ("", None)
            ]
            output: dict[str, Any] = {
                "size_group": group,
                "major_axis_range": ellipse_group_label(
                    group, small_boundary_px, large_boundary_px),
                "metric": display_name,
                "metric_key": metric,
                "unit": unit,
                "direction": direction,
                "count": len(values),
            }
            if values:
                array = np.asarray(values, dtype=np.float64)
                output.update(
                    mean=round(float(np.mean(array)), 6),
                    std=round(float(np.std(array)), 6),
                    min=round(float(np.min(array)), 6),
                    p05=round(float(np.percentile(array, 5)), 6),
                    median=round(float(np.median(array)), 6),
                    p95=round(float(np.percentile(array, 95)), 6),
                    max=round(float(np.max(array)), 6),
                )
                if metric in ERROR_COLUMNS:
                    pass_count = int(np.count_nonzero(array <= threshold_px))
                    output.update(
                        threshold=threshold_px,
                        pass_count=pass_count,
                        pass_rate_percent=round(
                            pass_count / len(values) * 100.0, 6),
                    )
            output_rows.append(output)

    write_metric_summary_rows(path, output_rows)
    return output_rows


def draw_all_metrics_by_outer_size_chart(
    path: Path,
    metric_rows: list[dict[str, Any]],
    small_boundary_px: float,
    large_boundary_px: float,
    metric_info: list[tuple[str, str, str, str]] | None = None,
    title: str = "Metrics by Fitted Outer-Ellipse Major Axis",
) -> None:
    """将全部指标的均值做成总体/小/中/大四列看板。"""
    selected_metric_info = (
        GROUPED_METRIC_INFO if metric_info is None else metric_info
    )
    groups = ("overall", "small", "medium", "large")
    width = 1680
    row_height = 58
    top = 105
    left = 430
    height = top + row_height * len(selected_metric_info) + 70
    canvas = np.full((height, width, 3), 248, dtype=np.uint8)
    cv2.putText(
        canvas, title,
        (80, 54), cv2.FONT_HERSHEY_SIMPLEX, 1.08,
        (25, 25, 25), 2, cv2.LINE_AA)
    cv2.putText(
        canvas, "Cell value = mean (sample count)",
        (80, 86), cv2.FONT_HERSHEY_SIMPLEX, 0.58,
        (70, 70, 70), 1, cv2.LINE_AA)

    column_width = (width - left - 40) // len(groups)
    headers = (
        "OVERALL",
        f"SMALL <{small_boundary_px:g}",
        f"MEDIUM {small_boundary_px:g}-{large_boundary_px:g}",
        f"LARGE >{large_boundary_px:g}",
    )
    for index, header in enumerate(headers):
        x = left + index * column_width
        cv2.rectangle(canvas, (x, top - 42),
                      (x + column_width, top), (215, 225, 235), -1)
        cv2.putText(canvas, header, (x + 18, top - 14),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.58,
                    (30, 30, 30), 2, cv2.LINE_AA)

    lookup = {
        (str(row["metric_key"]), str(row["size_group"])): row
        for row in metric_rows
    }
    for row_index, (metric, display_name, unit, _) in enumerate(
        selected_metric_info
    ):
        y1 = top + row_index * row_height
        y2 = y1 + row_height
        background = (238, 238, 238) if row_index % 2 == 0 else (250, 250, 250)
        cv2.rectangle(canvas, (40, y1), (width - 40, y2), background, -1)
        cv2.putText(canvas, display_name, (55, y1 + 35),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.48,
                    (35, 35, 35), 1, cv2.LINE_AA)
        for group_index, group in enumerate(groups):
            x = left + group_index * column_width
            cv2.line(canvas, (x, y1), (x, y2), (205, 205, 205), 1)
            item = lookup[(metric, group)]
            count = int(item["count"])
            if count == 0:
                text = "-- (n=0)"
            else:
                mean = float(item["mean"])
                text = (
                    f"{mean * 100.0:.2f}% (n={count})"
                    if unit == "ratio"
                    else f"{mean:.3f}px (n={count})"
                )
            cv2.putText(canvas, text, (x + 18, y1 + 35),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.48,
                        (30, 30, 30), 1, cv2.LINE_AA)
    cv2.rectangle(canvas, (40, top - 42), (width - 40, height - 70),
                  (90, 90, 90), 2)
    write_image(path, canvas)


def draw_ellipse_iou_group_chart(
    path: Path,
    summary_rows: list[dict[str, Any]],
) -> None:
    """绘制总体及三档 IoU：均值柱、最小/最大范围和 P95 标记。"""
    width, height = 1280, 860
    canvas = np.full((height, width, 3), 248, dtype=np.uint8)
    left, right, top, bottom = 120, width - 70, 110, 600
    cv2.rectangle(canvas, (left, top), (right, bottom), (70, 70, 70), 2)
    for tick in range(0, 101, 10):
        y = bottom - int((bottom - top) * tick / 100.0)
        cv2.line(canvas, (left, y), (right, y), (215, 215, 215), 1)
        cv2.putText(canvas, f"{tick}%", (55, y + 6),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.58, (60, 60, 60), 1,
                    cv2.LINE_AA)

    cv2.putText(canvas, "Outer Label vs Fitted Ellipse IoU by Major Axis",
                (160, 55), cv2.FONT_HERSHEY_SIMPLEX, 1.05,
                (20, 20, 20), 2, cv2.LINE_AA)
    colors = [
        (150, 120, 210),
        (90, 170, 255),
        (80, 190, 80),
        (210, 130, 70),
    ]
    valid_count = (
        int(summary_rows[0]["count"]) if summary_rows else 0
    )
    slot_width = (right - left) / max(1, len(summary_rows))
    for index, (row, color) in enumerate(zip(summary_rows, colors)):
        center_x = int(left + slot_width * (index + 0.5))
        count = int(row["count"])
        mean_percent = float(row.get("mean_iou_percent", 0.0) or 0.0)
        min_percent = float(row.get("min_iou", 0.0) or 0.0) * 100.0
        max_percent = float(row.get("max_iou", 0.0) or 0.0) * 100.0
        p95_percent = float(row.get("p95_iou", 0.0) or 0.0) * 100.0
        bar_width = int(slot_width * 0.42)
        bar_top = bottom - int((bottom - top) * mean_percent / 100.0)
        cv2.rectangle(canvas, (center_x - bar_width // 2, bar_top),
                      (center_x + bar_width // 2, bottom), color, -1)
        min_y = bottom - int((bottom - top) * min_percent / 100.0)
        max_y = bottom - int((bottom - top) * max_percent / 100.0)
        whisker_x = center_x + bar_width // 2 + 18
        cv2.line(canvas, (whisker_x, min_y), (whisker_x, max_y),
                 (70, 70, 70), 2)
        cv2.line(canvas, (whisker_x - 10, min_y), (whisker_x + 10, min_y),
                 (70, 70, 70), 2)
        cv2.line(canvas, (whisker_x - 10, max_y), (whisker_x + 10, max_y),
                 (70, 70, 70), 2)
        p95_y = bottom - int((bottom - top) * p95_percent / 100.0)
        cv2.line(canvas, (center_x - bar_width // 2 - 10, p95_y),
                 (center_x + bar_width // 2 + 10, p95_y),
                 (0, 120, 255), 3)
        cv2.putText(canvas, f"mean {mean_percent:.2f}%",
                    (center_x - 90, max(top + 24, bar_top - 18)),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.62, (30, 30, 30), 2,
                    cv2.LINE_AA)
        proportion = count / valid_count * 100.0 if valid_count else 0.0
        cv2.putText(canvas, f"P95 {p95_percent:.2f}% | n={count}",
                    (center_x - 125, bottom + 38),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.54, (40, 40, 40), 1,
                    cv2.LINE_AA)
        cv2.putText(
            canvas,
            f"min {min_percent:.2f}% | max {max_percent:.2f}%",
            (center_x - 125, bottom + 72),
            cv2.FONT_HERSHEY_SIMPLEX, 0.48, (40, 40, 40), 1,
            cv2.LINE_AA,
        )
        cv2.putText(
            canvas,
            f"valid share {proportion:.2f}%",
            (center_x - 125, bottom + 106),
            cv2.FONT_HERSHEY_SIMPLEX, 0.48, (40, 40, 40), 1,
            cv2.LINE_AA,
        )
        range_text = str(row["major_axis_range"])
        cv2.putText(canvas, range_text,
                    (center_x - 155, bottom + 140),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.47, (40, 40, 40), 1,
                    cv2.LINE_AA)
    cv2.putText(
                canvas,
                "Bar: mean IoU | Orange line: P95 | Whisker: min-max",
                (left, height - 28), cv2.FONT_HERSHEY_SIMPLEX, 0.58,
                (40, 40, 40), 1, cv2.LINE_AA)
    write_image(path, canvas)


def lowest_ellipse_iou_records(
    rows: list[dict[str, Any]],
    top_k: int,
    output_dir: Path,
) -> list[dict[str, Any]]:
    """生成最低 IoU 排行，Excel 和 TXT 回退共用。"""
    ranked = sorted(
        (row for row in rows if row.get("outer_ellipse_status") == "ok"),
        key=lambda row: (
            float(row["outer_label_ellipse_iou"]),
            str(row["image"]).lower(),
        ),
    )[:top_k]
    return [
        {
            "rank": rank,
            "iou": float(row["outer_label_ellipse_iou"]),
            "major_axis_px": float(row["outer_ellipse_major_axis_px"]),
            "size_group": row["outer_size_group"],
            "source": row["outer_ellipse_source"],
            "image": str(row["image"]),
            "visualization": str(output_dir / f"rank_{rank:02d}.jpg"),
        }
        for rank, row in enumerate(ranked, start=1)
    ]


def write_lowest_ellipse_iou_visualizations(
    output_dir: Path,
    rows: list[dict[str, Any]],
    images_path: Path,
    visualization_dir: Path,
    top_k: int,
) -> list[dict[str, Any]]:
    """保存 IoU 最低的外圈椭圆样本，便于快速定位拟合异常。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    records = lowest_ellipse_iou_records(rows, top_k, output_dir)
    for record in records:
        rank = int(record["rank"])
        source_image = Path(str(record["image"]))
        source_visualization = visualization_path_for_image(
            source_image, images_path, visualization_dir)
        visualized = read_image(source_visualization)
        if visualized is None:
            continue
        overlay = visualized.copy()
        cv2.rectangle(overlay, (6, 6), (min(850, overlay.shape[1] - 6), 82),
                      (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.72, visualized, 0.28, 0.0, visualized)
        iou = float(record["iou"])
        cv2.putText(
            visualized,
            f"LOWEST OUTER ELLIPSE IoU RANK {rank:02d}/{top_k:02d}: "
            f"{iou:.4f} ({iou * 100.0:.2f}%)",
            (16, 38), cv2.FONT_HERSHEY_SIMPLEX, 0.62,
            (0, 165, 255), 2, cv2.LINE_AA)
        cv2.putText(
            visualized,
            f"major={float(record['major_axis_px']):.2f}px | "
            f"group={record['size_group']} | source={record['source']}",
            (16, 69), cv2.FONT_HERSHEY_SIMPLEX, 0.52,
            (255, 255, 255), 1, cv2.LINE_AA)
        write_image(Path(str(record["visualization"])), visualized)
    return records


def draw_status_visualization(image: np.ndarray, status_text: str) -> np.ndarray:
    """为无法进入误差计算的图片也生成一张明确的检测状态图。"""
    canvas = image.copy()
    overlay = canvas.copy()
    panel_width = min(max(220, canvas.shape[1] - 12), 900)
    cv2.rectangle(overlay, (6, 6), (panel_width, 62), (0, 0, 0), -1)
    cv2.addWeighted(overlay, 0.72, canvas, 0.28, 0.0, canvas)
    cv2.putText(
        canvas,
        f"DETECTION EVALUATION: {status_text}",
        (16, 43),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.66,
        (0, 165, 255),
        2,
        cv2.LINE_AA,
    )
    return canvas


def main() -> int:
    args = parse_args()

    if not args.model.is_file():
        print(f"模型文件不存在：{args.model}", file=sys.stderr)
        return 2
    if not args.labels.is_dir():
        print(f"标签目录不存在：{args.labels}", file=sys.stderr)
        return 2

    try:
        image_paths = list_images(args.images, args.recursive)
    except (FileNotFoundError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2
    if not image_paths:
        print(f"没有找到支持的图片：{args.images}", file=sys.stderr)
        return 2

    detection_dir = args.output / "detection"
    ellipse_iou_dir = args.output / "ellipse_iou"
    visualization_dir = detection_dir / "visualizations"
    over_threshold_dir = detection_dir / "over_threshold"
    top_k_dir = detection_dir / "top10_max_error"
    ellipse_iou_visualization_dir = ellipse_iou_dir / "visualizations"
    lowest_ellipse_iou_dir = ellipse_iou_dir / "lowest_50"
    detection_dir.mkdir(parents=True, exist_ok=True)
    ellipse_iou_dir.mkdir(parents=True, exist_ok=True)
    if args.save_vis:
        visualization_dir.mkdir(parents=True, exist_ok=True)
        over_threshold_dir.mkdir(parents=True, exist_ok=True)
        ellipse_iou_visualization_dir.mkdir(parents=True, exist_ok=True)

    print(f"加载模型：{args.model}")
    label_scope = (
        "全部标签类别"
        if args.class_id is None
        else f"仅标签类别 {args.class_id}"
    )
    prediction_scope = (
        "预测类别不限"
        if args.ignore_pred_class
        else "标签与预测必须同类别"
    )
    print(
        f"匹配规则：{label_scope}，{prediction_scope}，"
        f"按 {args.selection.upper()} 选择最佳组合"
    )
    print(
        f"外圈椭圆：cls{args.outer_class_id}，"
        f"长轴分组 <{args.small_major_axis_px:g}px / "
        f"{args.small_major_axis_px:g}–{args.large_major_axis_px:g}px / "
        f">{args.large_major_axis_px:g}px"
    )
    model = YOLO(str(args.model))
    rows: list[dict[str, Any]] = []

    for index, image_path in enumerate(image_paths, start=1):
        label_path = label_path_for_image(image_path, args.images, args.labels)
        row = empty_row(image_path, label_path)
        print(f"[{index}/{len(image_paths)}] {image_path.name}")

        image = read_image(image_path)
        if image is None:
            row["status"] = "image_read_error"
            row["message"] = "无法读取图片"
            rows.append(row)
            continue

        image_height, image_width = image.shape[:2]
        try:
            all_gt_targets = read_segmentation_labels(
                label_path,
                image_width,
                image_height,
                requested_class=None,
            )
            gt_targets = (
                all_gt_targets
                if args.class_id is None
                else [
                    target for target in all_gt_targets
                    if target[0] == args.class_id
                ]
            )
            if not gt_targets:
                raise ValueError(
                    f"标签中不存在类别 {args.class_id}：{label_path}")
        except FileNotFoundError as exc:
            row["status"] = "missing_label"
            row["message"] = str(exc)
            row["outer_ellipse_status"] = "missing_label"
            if args.save_vis:
                write_image(
                    visualization_path_for_image(
                        image_path, args.images, visualization_dir),
                    draw_status_visualization(image, "MISSING LABEL"),
                )
                clean = draw_outer_ellipse_iou_visualization(
                    image, {"status": "missing_label"})
                write_image(
                    visualization_path_for_image(
                        image_path, args.images,
                        ellipse_iou_visualization_dir),
                    clean,
                )
            rows.append(row)
            continue
        except ValueError as exc:
            row["status"] = "invalid_label"
            row["message"] = str(exc)
            row["outer_ellipse_status"] = "invalid_label"
            if args.save_vis:
                write_image(
                    visualization_path_for_image(
                        image_path, args.images, visualization_dir),
                    draw_status_visualization(image, "INVALID LABEL"),
                )
                clean = draw_outer_ellipse_iou_visualization(
                    image, {"status": "invalid_label"})
                write_image(
                    visualization_path_for_image(
                        image_path, args.images,
                        ellipse_iou_visualization_dir),
                    clean,
                )
            rows.append(row)
            continue

        gt_class, gt_bbox, polygon = largest_ground_truth(gt_targets)
        row["gt_target_count"] = len(gt_targets)
        row["match_mode"] = (
            f"max_{args.selection}_all_classes"
            if args.ignore_pred_class
            else f"max_{args.selection}_same_class"
        )

        predict_kwargs: dict[str, Any] = {
            "source": image,
            "conf": args.conf,
            "imgsz": args.imgsz,
            "retina_masks": True,
            "verbose": False,
        }
        if args.device is not None:
            predict_kwargs["device"] = args.device

        try:
            result = model.predict(**predict_kwargs)[0]
        except Exception as exc:  # Ultralytics errors vary by backend/version.
            row["status"] = "inference_error"
            row["message"] = str(exc)
            row["gt_class"] = gt_class
            row["outer_ellipse_status"] = "inference_error"
            if args.save_vis:
                clean = draw_outer_ellipse_iou_visualization(
                    image, {"status": "inference_error"})
                write_image(
                    visualization_path_for_image(
                        image_path, args.images,
                        ellipse_iou_visualization_dir),
                    clean,
                )
                visualized = draw_visualization(
                    image,
                    polygon,
                    gt_bbox,
                    pred_bbox=None,
                    errors=None,
                    confidence=None,
                    gt_class=gt_class,
                    status_text="INFERENCE ERROR",
                )
                write_image(
                    visualization_path_for_image(
                        image_path,
                        args.images,
                        visualization_dir,
                    ),
                    visualized,
                )
            rows.append(row)
            continue

        # 外圈椭圆评估与原有检测框误差评估相互独立。
        # 即使某张图的椭圆拟合异常，也不中断原有七项误差统计。
        try:
            outer_evaluation = evaluate_outer_ellipse(
                image, result, all_gt_targets, args)
        except Exception as exc:
            outer_evaluation = {
                "status": "outer_ellipse_error",
                "error": str(exc),
            }
            print(f"  [WARNING] 外圈椭圆评估失败：{exc}", file=sys.stderr)
        update_row_with_outer_ellipse(row, outer_evaluation)
        if args.save_vis:
            ellipse_iou_visualization = (
                draw_outer_ellipse_iou_visualization(
                    image, outer_evaluation)
            )
            ellipse_iou_path = visualization_path_for_image(
                image_path,
                args.images,
                ellipse_iou_visualization_dir,
            )
            write_image(ellipse_iou_path, ellipse_iou_visualization)

        (
            gt_class,
            gt_bbox,
            polygon,
            pred_bbox,
            pred_class,
            confidence,
            prediction_count,
            candidate_pair_count,
            matched_iou,
        ) = choose_best_match(
            result,
            gt_targets,
            args.ignore_pred_class,
            args.selection,
            image_width,
            image_height,
        )

        row["gt_class"] = gt_class
        row["prediction_count"] = prediction_count
        row["candidate_pair_count"] = candidate_pair_count
        row["threshold_px"] = rounded(args.threshold)
        row["gt_x1"], row["gt_y1"], row["gt_x2"], row["gt_y2"] = (
            rounded(value) for value in gt_bbox
        )

        if pred_bbox is None:
            row["status"] = "no_detection"
            row["message"] = (
                "模型没有输出检测框"
                if prediction_count == 0
                else "类别约束下没有可用的标签框与预测框组合"
            )
            if args.save_vis:
                visualized = draw_visualization(
                    image,
                    polygon,
                    gt_bbox,
                    pred_bbox=None,
                    errors=None,
                    confidence=None,
                    gt_class=gt_class,
                    status_text="NO DETECTION",
                )
                visualization_path = visualization_path_for_image(
                    image_path, args.images, visualization_dir)
                write_image(visualization_path, visualized)
            rows.append(row)
            continue

        errors = calculate_errors(gt_bbox, pred_bbox)
        row.update(
            {
                "status": "ok",
                "message": (
                    f"从 {len(gt_targets)} 个标签目标、"
                    f"{prediction_count} 个预测框组成的 "
                    f"{candidate_pair_count} 个候选组合中，"
                    f"按 {args.selection.upper()} 选择："
                    f"GT class {gt_class} <-> Pred class {pred_class}，"
                    f"IoU={float(matched_iou):.6f}"
                ),
                "pred_class": pred_class,
                "confidence": rounded(confidence),
                "pred_x1": rounded(pred_bbox[0]),
                "pred_y1": rounded(pred_bbox[1]),
                "pred_x2": rounded(pred_bbox[2]),
                "pred_y2": rounded(pred_bbox[3]),
            }
        )
        for key, value in errors.items():
            row[key] = value if isinstance(value, str) else rounded(value)

        pass_results = [
            float(errors[metric]) <= args.threshold for metric in ERROR_COLUMNS
        ]
        for pass_column, passed in zip(PASS_COLUMNS, pass_results):
            row[pass_column] = int(passed)
        row["all_metrics_within_threshold"] = int(all(pass_results))
        rows.append(row)

        if args.save_vis:
            visualized = draw_visualization(
                image,
                polygon,
                gt_bbox,
                pred_bbox,
                errors,
                float(confidence),
                gt_class=gt_class,
                pred_class=pred_class,
                threshold=args.threshold,
            )
            visualization_path = visualization_path_for_image(
                image_path, args.images, visualization_dir)
            write_image(visualization_path, visualized)
            if not all(pass_results):
                write_image(
                    visualization_path_for_image(
                        image_path,
                        args.images,
                        over_threshold_dir,
                    ),
                    visualized,
                )

    detection_excel_path = detection_dir / "detection_statistics.xlsx"
    ellipse_excel_path = ellipse_iou_dir / "ellipse_iou_statistics.xlsx"
    detection_errors_chart_path = detection_dir / "grouped_errors.png"
    ellipse_chart_path = ellipse_iou_dir / "overall_and_grouped.png"

    detection_error_rows = detection_metric_statistics(
        rows,
        args.threshold,
        args.small_major_axis_px,
        args.large_major_axis_px,
    )
    draw_all_metrics_by_outer_size_chart(
        detection_errors_chart_path,
        detection_error_rows,
        args.small_major_axis_px,
        args.large_major_axis_px,
        metric_info=DETECTION_ERROR_METRIC_INFO,
        title="Detection Errors by Fitted Outer-Ellipse Major Axis",
    )
    ellipse_summary_rows = ellipse_iou_group_statistics(
        rows, args.small_major_axis_px, args.large_major_axis_px
    )
    draw_ellipse_iou_group_chart(ellipse_chart_path, ellipse_summary_rows)

    top_records = detection_top_k_records(rows, args.top_k, top_k_dir)
    lowest_records = lowest_ellipse_iou_records(
        rows, args.lowest_iou_count, lowest_ellipse_iou_dir)
    top_k_count = 0
    if args.save_vis:
        top_records, top_k_count = write_overall_top_k_visualizations(
            top_k_dir,
            rows,
            args.images,
            visualization_dir,
            args.top_k,
        )
        lowest_records = write_lowest_ellipse_iou_visualizations(
            lowest_ellipse_iou_dir,
            rows,
            args.images,
            ellipse_iou_visualization_dir,
            args.lowest_iou_count,
        )

    excel_written = False
    if HAS_OPENPYXL:
        try:
            write_detection_excel(
                detection_excel_path,
                detection_error_rows,
                rows,
                top_records,
            )
            write_ellipse_iou_excel(
                ellipse_excel_path,
                ellipse_summary_rows,
                rows,
                lowest_records,
            )
            excel_written = True
        except Exception as exc:
            print(
                f"[WARNING] Excel 写入失败，将自动回退 TXT：{exc}",
                file=sys.stderr,
            )

    if excel_written:
        statistics_mode = "Excel"
        detection_statistics_paths = [detection_excel_path]
        ellipse_statistics_paths = [ellipse_excel_path]
    else:
        # 服务器没有 openpyxl 时自动回退 TXT，不中止当前批量评估。
        detection_overall_path = detection_dir / "overall_errors.txt"
        detection_grouped_path = detection_dir / "grouped_errors.txt"
        per_image_path = detection_dir / "per_image_errors.txt"
        ellipse_summary_path = ellipse_iou_dir / "overall_and_grouped.txt"
        write_detection_statistics_txt(
            detection_overall_path,
            detection_error_rows,
            ("overall",),
            args.threshold,
        )
        write_detection_statistics_txt(
            detection_grouped_path,
            detection_error_rows,
            ("small", "medium", "large"),
            args.threshold,
        )
        write_per_image_detection_txt(per_image_path, rows)
        write_ellipse_iou_statistics_txt(
            ellipse_summary_path,
            rows,
            args.small_major_axis_px,
            args.large_major_axis_px,
        )
        write_ranking_txt(
            top_k_dir / "top10.txt",
            ["排名", "图片", "单图最大误差(px)", "最大误差指标", "可视化文件"],
            (
                [
                    record["rank"], record["image"], record["max_error_px"],
                    record["metric"], record["visualization"],
                ]
                for record in top_records
            ),
        )
        write_ranking_txt(
            lowest_ellipse_iou_dir / "lowest_iou.txt",
            [
                "排名", "IoU", "IoU占比(%)", "外圈长轴(px)", "分组",
                "拟合来源", "原图片", "可视化文件",
            ],
            (
                [
                    record["rank"], record["iou"],
                    float(record["iou"]) * 100.0,
                    record["major_axis_px"], record["size_group"],
                    record["source"], record["image"],
                    record["visualization"],
                ]
                for record in lowest_records
            ),
        )
        statistics_mode = "TXT（Excel 不可用或写入失败，已自动回退）"
        detection_statistics_paths = [
            detection_overall_path, detection_grouped_path, per_image_path]
        ellipse_statistics_paths = [ellipse_summary_path]

    status_counts = Counter(row["status"] for row in rows)
    outer_status_counts = Counter(
        str(row.get("outer_ellipse_status") or "not_evaluated")
        for row in rows
    )
    print("\n处理完成")
    print(f"统计表格式：{statistics_mode}")
    print(
        "检测统计："
        + "；".join(str(path) for path in detection_statistics_paths)
    )
    print(f"检测分组看板：{detection_errors_chart_path}")
    print(
        "外圈椭圆 IoU 统计："
        + "；".join(str(path) for path in ellipse_statistics_paths)
    )
    print(f"外圈椭圆 IoU 统计图：{ellipse_chart_path}")
    if args.save_vis:
        print(f"全部检测可视化：{visualization_dir}")
        print(f"全部椭圆 IoU 可视化：{ellipse_iou_visualization_dir}")
        over_threshold_count = sum(
            row["status"] == "ok"
            and not bool(int(row["all_metrics_within_threshold"]))
            for row in rows
        )
        print(
            f"任一误差超过 {args.threshold:g} px 的图片："
            f"{over_threshold_dir}（{over_threshold_count} 张）"
        )
        print(
            f"总体最大误差 Top {args.top_k}："
            f"{top_k_dir}（实际 {top_k_count} 张）"
        )
        print(
            f"外圈椭圆 IoU 最低 {args.lowest_iou_count}："
            f"{lowest_ellipse_iou_dir}"
        )
    print("状态统计：")
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count}")

    valid_rows = [row for row in rows if row["status"] == "ok"]
    if valid_rows:
        all_metrics_pass_count = sum(
            int(row["all_metrics_within_threshold"]) for row in valid_rows
        )
        all_metrics_pass_rate = (
            all_metrics_pass_count / len(valid_rows) * 100.0
        )
        print(
            f"七项误差全部 ≤ {args.threshold:g} px 的整体达标率："
            f"{all_metrics_pass_rate:.2f}% "
            f"({all_metrics_pass_count}/{len(valid_rows)})"
        )
    print("外圈椭圆 IoU 总体及三档统计：")
    for group_row in ellipse_summary_rows:
        if int(group_row["count"]) == 0:
            print(f"  {group_row['major_axis_range']}: 0 张")
            continue
        print(
            f"  {group_row['major_axis_range']}: "
            f"mean={float(group_row['mean_iou_percent']):.2f}% "
            f"median={float(group_row['median_iou']) * 100.0:.2f}% "
            f"n={group_row['count']}"
        )
    print("外圈椭圆状态：")
    for status, count in sorted(outer_status_counts.items()):
        print(f"  {status}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
